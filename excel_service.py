import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def carregar_excel(excel_bytes):
    """
    Reads an Excel file and returns sanitized dataframe and suggested column mappings.
    """
    df = pd.read_excel(io.BytesIO(excel_bytes))
    
    # Strip whitespace from column names
    df.columns = [str(c).strip() for c in df.columns]
    
    # Drop rows that are completely blank
    df = df.dropna(how='all').reset_index(drop=True)
    
    cols = list(df.columns)
    
    col_data = None
    col_valor = None
    col_desc = None
    
    for c in cols:
        c_lower = str(c).lower()
        if not col_data and any(k in c_lower for k in ['data', 'dt', 'vencimento', 'pagamento']):
            col_data = c
        elif not col_valor and any(k in c_lower for k in ['valor', 'val', 'quantia', 'preco', 'preço', 'montante']):
            col_valor = c
        elif not col_desc and any(k in c_lower for k in ['desc', 'nome', 'cod', 'código', 'ref', 'item']):
            col_desc = c
            
    if not col_data and len(cols) >= 1:
        col_data = cols[0]
    if not col_valor and len(cols) >= 2:
        col_valor = cols[1]
        
    return df, {
        "col_data": col_data,
        "col_valor": col_valor,
        "col_desc": col_desc
    }

def gerar_excel_atualizado(df_original, resultados, col_data, col_valor, col_desc=None):
    """
    Updates the DataFrame with BCB calculated columns and exports a formatted Excel file.
    """
    df = df_original.copy()
    
    datas_finais = []
    valores_corrigidos = []
    diferencas = []
    fatores = []
    percentuais = []
    statuses = []
    
    for i, r in enumerate(resultados):
        datas_finais.append(r.get('data_final', ''))
        v_corr = r.get('valor_corrigido_num')
        v_orig = r.get('valor_original_num', 0)
        
        if v_corr is not None:
            valores_corrigidos.append(v_corr)
            diferencas.append(round(v_corr - v_orig, 2))
        else:
            valores_corrigidos.append(None)
            diferencas.append(None)
            
        fatores.append(r.get('fator_correcao', '-'))
        percentuais.append(r.get('percentual', '-'))
        statuses.append(r.get('status', 'Pendente'))
        
    df['Data Final (Correção)'] = datas_finais
    df['Valor Corrigido (R$)'] = valores_corrigidos
    df['Diferença / Rendimento (R$)'] = diferencas
    df['Fator de Correção (BCB)'] = fatores
    df['Percentual (%)'] = percentuais
    df['Status'] = statuses

    # Save formatted Excel with openpyxl
    out_buf = io.BytesIO()
    with pd.ExcelWriter(out_buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Pagamentos Corrigidos')
        wb = writer.book
        ws = writer.sheets['Pagamentos Corrigidos']
        
        # Styles
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                col_header = str(ws.cell(row=1, column=col_idx).value)
                
                if "Valor" in col_header or "Diferença" in col_header:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif "Data" in col_header or "Status" in col_header:
                    cell.alignment = Alignment(horizontal="center")
                    
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    return out_buf.getvalue()
